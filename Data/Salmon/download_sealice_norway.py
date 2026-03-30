"""
Download weekly national sea lice + temperature statistics from BarentsWatch Fish Health API
Output: Excel file with one row per week from 2012 to today

Requirements:
    pip install requests openpyxl pandas

Setup:
    1. Register at https://www.barentswatch.no/minside/
    2. Go to "For developers" and create an API client (client_credentials flow)
    3. Fill in CLIENT_ID and CLIENT_SECRET below, then run:
           python download_sealice_norway.py
"""

import requests
import pandas as pd
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration — fill these in ─────────────────────────────────────────────
CLIENT_ID     = "YOUR_CLIENT_ID"
CLIENT_SECRET = "YOUR_CLIENT_SECRET"
OUTPUT_FILE   = "sealice_norway_weekly.xlsx"
# ──────────────────────────────────────────────────────────────────────────────

TOKEN_URL  = "https://id.barentswatch.no/connect/token"
API_BASE   = "https://www.barentswatch.no/bwapi"
START_YEAR, START_WEEK = 2012, 1


def get_token():
    resp = requests.post(
        TOKEN_URL,
        data={
            "client_id":     CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "grant_type":    "client_credentials",
            "scope":         "api",
        },
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def fetch_week(token, year, week):
    url  = f"{API_BASE}/v1/geodata/fishhealth/locality/week/{year}/{week}"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    if resp.status_code == 204:
        return []
    resp.raise_for_status()
    return resp.json()


def iter_yearweeks(start_year, start_week):
    today = date.today()
    end_year, end_week = today.isocalendar()[0], today.isocalendar()[1]
    y, w = start_year, start_week
    while (y, w) <= (end_year, end_week):
        yield y, w
        d   = date.fromisocalendar(y, w, 1) + timedelta(weeks=1)
        iso = d.isocalendar()
        y, w = iso[0], iso[1]


def safe_mean(values):
    clean = [v for v in values if v is not None]
    return round(sum(clean) / len(clean), 4) if clean else None


def aggregate_week(year, week, localities):
    reported = [l for l in localities if l.get("hasReported")]
    if not reported:
        return None

    adult_female  = safe_mean([l.get("avgAdultFemaleLice")  for l in reported])
    sea_temp      = safe_mean([l.get("avgSeaTemperature")   for l in reported])
    n             = len(reported)
    above_limit   = round(sum(1 for l in reported if l.get("isAboveLimit")) / n * 100, 2) if n else None
    any_treatment = round(sum(
        1 for l in reported if any([
            l.get("hasBathTreatment"),
            l.get("hasMechanicalTreatment"),
            l.get("hasInFeedTreatment"),
        ])
    ) / n * 100, 2) if n else None

    return {
        "year":                 year,
        "week":                 week,
        "iso_week":             f"{year}-W{week:02d}",
        "localities_reporting": n,
        "avg_adult_female_lice": adult_female,
        "avg_sea_temp_3m":       sea_temp,
        "pct_above_limit":       above_limit,
        "pct_any_treatment":     any_treatment,
    }


def write_excel(rows, path):
    df = pd.DataFrame(rows)

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Weekly Norway"

    # Styles
    DARK_BLUE   = "1F4E79"
    LIGHT_BLUE  = "D6E4F0"
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", start_color=DARK_BLUE)
    alt_fill    = PatternFill("solid", start_color=LIGHT_BLUE)
    cell_font   = Font(name="Arial", size=10)
    center      = Alignment(horizontal="center")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )

    columns = [
        ("Year",                    "year",                  10),
        ("Week",                    "week",                   8),
        ("ISO Week",                "iso_week",              12),
        ("Localities Reporting",    "localities_reporting",  22),
        ("Avg Adult Female Lice",   "avg_adult_female_lice", 22),
        ("Avg Sea Temp 3m (°C)",    "avg_sea_temp_3m",       22),
        ("% Above Limit (0.5)",     "pct_above_limit",       20),
        ("% Any Treatment",         "pct_any_treatment",     20),
    ]

    # Header row
    for col_idx, (header, _, width) in enumerate(columns, 1):
        cell            = ws.cell(row=1, column=col_idx, value=header)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, (_, field, _) in enumerate(columns, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=getattr(row, field))
            cell.font      = cell_font
            cell.alignment = center
            cell.border    = thin_border
            if fill:
                cell.fill = fill

    # Number formatting
    pct_cols  = [7, 8]   # % Above Limit, % Any Treatment
    temp_cols = [6]      # Avg Sea Temp
    lice_cols = [5]      # Avg Adult Female Lice

    for row_idx in range(2, len(df) + 2):
        for c in pct_cols:
            ws.cell(row=row_idx, column=c).number_format = '0.00"%"'
        for c in temp_cols:
            ws.cell(row=row_idx, column=c).number_format = '0.00'
        for c in lice_cols:
            ws.cell(row=row_idx, column=c).number_format = '0.0000'

    # Summary stats block to the right
    ws.cell(row=1,  column=10, value="Summary Statistics").font = Font(bold=True, name="Arial", size=10)
    stats = [
        ("Total weeks",           f"=COUNTA(C2:C{len(df)+1})"),
        ("First week",            f"=MIN(C2:C{len(df)+1})"),
        ("Last week",             f"=MAX(C2:C{len(df)+1})"),
        ("Avg lice (all time)",   f"=AVERAGE(E2:E{len(df)+1})"),
        ("Max lice (all time)",   f"=MAX(E2:E{len(df)+1})"),
        ("Avg temp (all time)",   f"=AVERAGE(F2:F{len(df)+1})"),
        ("Avg % above limit",     f"=AVERAGE(G2:G{len(df)+1})"),
        ("Avg % any treatment",   f"=AVERAGE(H2:H{len(df)+1})"),
    ]
    for i, (label, formula) in enumerate(stats, 2):
        ws.cell(row=i, column=10, value=label).font  = Font(bold=True, name="Arial", size=10)
        ws.cell(row=i, column=11, value=formula).font = Font(name="Arial", size=10)
    ws.column_dimensions["J"].width = 24
    ws.column_dimensions["K"].width = 16

    wb.save(path)


def main():
    print("Authenticating …")
    token = get_token()

    weeks     = list(iter_yearweeks(START_YEAR, START_WEEK))
    total     = len(weeks)
    rows      = []
    skipped   = 0

    print(f"Fetching {total} weeks (2012-W01 → today) — this takes a few minutes …\n")

    for i, (year, week) in enumerate(weeks, 1):
        if i % 50 == 0 or i == total:
            print(f"  {i}/{total} weeks …")

        localities = fetch_week(token, year, week)
        result     = aggregate_week(year, week, localities)

        if result:
            rows.append(result)
        else:
            skipped += 1

    print(f"\nAggregated {len(rows)} weeks ({skipped} weeks with no data skipped)")
    print(f"Writing to '{OUTPUT_FILE}' …")

    write_excel(rows, OUTPUT_FILE)

    print(f"\n✓ Done! Saved → '{OUTPUT_FILE}'")
    print(f"  Rows: {len(rows)}")
    print(f"  Columns: year, week, iso_week, localities_reporting,")
    print(f"           avg_adult_female_lice, avg_sea_temp_3m,")
    print(f"           pct_above_limit, pct_any_treatment")


if __name__ == "__main__":
    main()
